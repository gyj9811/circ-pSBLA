from keras.models import Sequential
from keras.optimizers import Adam
import tensorflow as tf
from keras.layers import LSTM
from keras.layers import Bidirectional, AveragePooling1D
from keras.layers import Convolution1D, MaxPooling1D, BatchNormalization
from keras.layers import Dense, Dropout, Flatten
from keras.layers import Activation
import keras.models
from keras.callbacks import EarlyStopping
from sklearn.model_selection import KFold
import argparse
import catboost
import openpyxl
import pandas
from gyjcode.onehotcrip import *
from gyjcode.attention import *

def cnnattention(parser, trainX, trainy, testX):
    protein = parser.protein
    batch_size = parser.batch_size
    n_epochs = parser.n_epochs
    nbfilter = parser.nbfilter
    hiddensize = parser.hiddensize
    predictions = []
    trainXeval = trainX
    trainYeval = trainy
    test_X = testX
    print(len(test_X))

    #cnn+attention
    kf = KFold(n_splits=5)
    for train_index, eval_index in kf.split(trainYeval):
        train_X = trainXeval[train_index]
        train_y = trainYeval[train_index]
        eval_X = trainXeval[eval_index]
        eval_y = trainYeval[eval_index]
        print('configure cnn network')
        modelcnn = Sequential()
        modelcnn.add(
            Convolution1D(input_shape=[101,4], filters=nbfilter, kernel_size=7, padding="valid",
                         activation="relu", strides=1))
        # modelcnn.add(
        #      Convolution1D( filters=nbfilter, kernel_size=7, padding="valid",
        #                   activation="relu", strides=1))
        modelcnn.add(BatchNormalization())
        modelcnn.add(Activation('relu'))
        modelcnn.add(AveragePooling1D(pool_size=5))
        #modelcnn.add(Dropout(0.5))
        modelcnn.add(Bidirectional(LSTM(hiddensize, return_sequences=True)))
        modelcnn.add(AttentionWithContext())
        modelcnn.add(Dense(nbfilter, activation='relu'))
        modelcnn.add(Dropout(0.25))
        modelcnn.add(Dense(2))
        modelcnn.add(Activation('softmax'))
        modelcnn.compile(loss='categorical_crossentropy', optimizer=Adam(lr=1e-4))
        print('model training')
        earlystopper = EarlyStopping(monitor='val_loss', patience=5, verbose=0)
        modelcnn.fit(train_X, train_y, batch_size=batch_size, nb_epoch=n_epochs, verbose=0, validation_data=(eval_X, eval_y),
                  callbacks=[earlystopper])
        cnnpredictions = modelcnn.predict_proba(test_X)[:, 1]
    return cnnpredictions

def cat(parser, trainX, trainy, trainy1, testX):
    protein = parser.protein
    batch_size = parser.batch_size
    n_epochs = parser.n_epochs
    hiddensize = parser.hiddensize
    nbfilter = parser.nbfilter
    trainXeval = trainX
    trainYeval = trainy
    trainYeval1 = trainy1
    test_X = testX

    kf = KFold(n_splits=5)
    for train_index, eval_index in kf.split(trainYeval):
        train_X = trainXeval[train_index]
        train_y = trainYeval[train_index]
        train_y1 = trainYeval1[train_index]
        eval_X = trainXeval[eval_index]
        eval_y = trainYeval[eval_index]
        eval_y1 = trainYeval1[eval_index]
        print('configure catboost')
        modelcnn = Sequential()
        modelcnn.add(
            Convolution1D(input_shape=[101,4], filters=nbfilter, kernel_size=7, padding="valid",
                         activation="relu", strides=1))
        # modelcnn.add(
        #     Convolution1D( filters=nbfilter, kernel_size=7, padding="valid",
        #                   activation="relu", strides=1))
        modelcnn.add(BatchNormalization())
        modelcnn.add(Activation('relu'))
        modelcnn.add(AveragePooling1D(pool_size=5))
        #modelcnn.add(Dropout(0.5))
        #modelcnn.add(BatchNormalization())
        modelcnn.add(Bidirectional(LSTM(hiddensize, return_sequences=True)))
        modelcnn.add(AttentionWithContext())
        modelcnn.add(Dense(nbfilter, activation='relu',  name='dense1'))
        modelcnn.add(Dropout(0.25))
        modelcnn.add(Dense(2, name='dense2'))
        modelcnn.compile(loss='categorical_crossentropy', optimizer=Adam(lr=1e-4))
        print('model training')
        earlystopper = EarlyStopping(monitor='val_loss', patience=5, verbose=0)
        modelcnn.fit(train_X, train_y, batch_size=batch_size, nb_epoch=n_epochs, verbose=0, validation_data=(eval_X, eval_y),
                     callbacks=[earlystopper])
        sub_model = keras.models.Model(inputs = modelcnn.input, outputs = modelcnn.get_layer('dense1').output)
        tr = sub_model.predict(train_X)
        te = sub_model.predict(test_X)
        va = sub_model.predict(eval_X)
    modelcat = catboost.CatBoostClassifier(iterations=2000,depth=6,learning_rate=0.01, loss_function='Logloss')
    print('model training')
    modelcat.fit(tr, train_y1, eval_set= (va, eval_y1), verbose=500)
    catpredictions = modelcat.predict_proba(te)[:,1]
    return catpredictions

def parse_arguments(parser):
    parser.add_argument('--protein', type=str, default='SFRS1',metavar='SFRS1')
    parser.add_argument('--nbfilter', type=int, default=102)
    parser.add_argument('--hiddensize', type=int, default=120)
    parser.add_argument('--batch_size', type=int, default=64)
    parser.add_argument('--n_epochs', type=int, default=30)
    args = parser.parse_args()
    return args

def getprotein(parser):
    protein = parser.protein
    return protein

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    args = parse_arguments(parse)
    protein = getprotein(args)
    trainXeval, test_X, trainYeval, test_y, trainYeval1 = dealwithdata(protein)
    prediction1 = cnnattention(args, trainXeval, trainYeval, test_X)
    prediction2 = cat(args, trainXeval, trainYeval, trainYeval1, test_X)
    test_y = test_y[:, 1]
    re_len = len(prediction1)

    #store cnn+attention
    f1 = openpyxl.Workbook()
    sheet1 = f1.active
    # 生成字段名（第一行）
    count = 0
    for i in range(re_len):
        sheet1.cell(row= count+1, column=0+1, value=str(test_y[count]))
        sheet1.cell(row= count+1, column=0+2, value=str(prediction1[count]))
        count += 1
    f1.save(protein + 'oneattention.xlsx')
    data1 = pandas.read_excel(protein+'oneattention.xlsx', index_col=0)
    data1.to_csv(protein+'oneattention.csv', encoding='utf-8')

    #store cnn+catboost
    f2 = openpyxl.Workbook()
    sheet1 = f2.active
    # 生成字段名（第一行）
    count = 0
    for i in range(re_len):
        sheet1.cell(row= count+1, column=0+1, value=str(test_y[count]))
        sheet1.cell(row= count+1, column=0+2, value=str(prediction2[count]))
        count += 1
    f2.save(protein + 'onecat.xlsx')
    data2 = pandas.read_excel(protein+'onecat.xlsx', index_col=0)
    data2.to_csv(protein+'onecat.csv', encoding='utf-8')

    f = openpyxl.Workbook()
    sheet1 = f.active
    predictions = []
    # 生成字段名（第一行）
    count = 0
    for i in range(re_len):
        prediction = (prediction1[count] + prediction2[count])/2
        predictions.append(prediction)
        sheet1.cell(row= count+1, column=0+1, value=str(test_y[count]))
        sheet1.cell(row= count+1, column=0+2, value=str(predictions[count]))
        count += 1
    f.save(protein + 'one.xlsx')
    data = pandas.read_excel(protein+'one.xlsx', index_col=0)
    data.to_csv(protein+'one.csv', encoding='utf-8')







